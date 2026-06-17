#!/usr/bin/env python3
"""
Конвертирует xlsx-файл в JSON для PWA «КИПиА».

Логика:
  1. Если xlsx-файл есть в репозитории (data/exam-tickets.xlsx) — берём его.
  2. Если нет — пробуем скачать с OneDrive по ссылке общего доступа.
  3. Конвертируем все листы в JSON и сохраняем.

Запускается через GitHub Actions (workflow_dispatch или schedule).
"""

import json
import os
import re
import sys

import openpyxl

# Попробуем импортировать requests (может не быть при оффлайн-конвертации)
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ============================================================
# Конфигурация
# ============================================================
ONEDRIVE_SHARE_URL = os.environ.get(
    "ONEDRIVE_SHARE_URL",
    "https://1drv.ms/x/c/c9414adb26fe5b28/IQDqsaT9uFR_T6HBWRyYhhEtAXDVcDPDoT3-O_um4E6V7O0?e=IKtgRl",
)

LOCAL_XLSX = os.environ.get("LOCAL_XLSX", "data/exam-tickets.xlsx")
OUTPUT_PATH = os.environ.get("OUTPUT_PATH", "data/exam-tickets.json")

SHEETS_CONFIG = {
    "4 разряд": {"id": "tickets-4", "title": "Билеты на 4 разряд"},
    "5 разряд": {"id": "tickets-5", "title": "Билеты на 5 разряд"},
    "6 разряд": {"id": "tickets-6", "title": "Билеты на 6 разряд"},
    "До 1000 В": {"id": "tickets-1000v", "title": "Билеты до 1000 В"},
}


# Префикс пути к картинкам в Excel, который нужно заменить на web-путь
_IMAGE_PATH_PREFIX = "ИОС\\2. рабочая документация\\по слесарям КИП\\инструктажи, проверка знаний\\Билеты на допуск к самостоятельной работе\\кип_app\\Билеты_Images\\"
# Префикс пути к файлам библиотеки в Excel
_LIBRARY_PATH_PREFIX = "Библиотека КИП и А\\"


def _is_http_url(s: str) -> bool:
    """Возвращает True, если строка начинается с http:// или https://.
    Такие значения — это рабочие ссылки (например, OneDrive share URL),
    их нельзя конвертировать в локальные пути.
    """
    return s.strip().lower().startswith(("http://", "https://"))


def _convert_image_path(raw_path: str) -> str:
    """Конвертирует локальный путь картинки из Excel в web-путь.

    Пример: ИОС\\...\\Билеты_Images\\приборы\\сх_термосопр.png → images/tickets/приборы/сх_термосопр.png

    OneDrive/HTTP-ссылки (https://1drv.ms/i/c/...) пропускаются как есть.
    """
    if not raw_path.strip():
        return ""
    # OneDrive/HTTP-ссылки не конвертируем — отдаём как есть
    if _is_http_url(raw_path):
        return raw_path.strip()
    # Заменяем обратные слеши на прямые
    p = raw_path.replace("\\", "/")
    # Ищем папку Билеты_Images и берём путь после неё
    marker = "Билеты_Images/"
    idx = p.find(marker)
    if idx >= 0:
        rel = p[idx + len(marker):]
        return "images/tickets/" + rel
    # Если маркер не найден, пробуем извлечь имя файла
    filename = p.split("/")[-1]
    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp')):
        return "images/tickets/" + filename
    return raw_path


def _convert_file_path(raw_path: str) -> str:
    """Конвертирует локальный путь файла из Excel в web-путь.

    Пример: Библиотека КИП и А\\Электробезопасность\\1. ПУЭ 6 и 7.pdf → library/Электробезопасность/1. ПУЭ 6 и 7.pdf

    OneDrive/HTTP-ссылки (https://1drv.ms/b/c/...) пропускаются как есть.
    """
    if not raw_path.strip():
        return ""
    # OneDrive/HTTP-ссылки не конвертируем — отдаём как есть
    if _is_http_url(raw_path):
        return raw_path.strip()
    p = raw_path.replace("\\", "/")
    # Убираем префикс «Библиотека КИП и А/»
    marker = "Библиотека КИП и А/"
    idx = p.find(marker)
    if idx >= 0:
        rel = p[idx + len(marker):]
        return "library/" + rel
    # Если путь начинается с ИОС — конвертируем как картинку
    if p.startswith("ИОС/"):
        return _convert_image_path(raw_path)
    return raw_path


def _unescape_url(url: str) -> str:
    """Расшифровывает экранированные символы в URL."""
    return url.replace("\\u0026", "&").replace("&amp;", "&")


def _try_download(dl_url: str, dest: str, headers: dict) -> bool:
    """Пытается скачать файл по URL и проверяет, что это xlsx (ZIP)."""
    try:
        r = requests.get(dl_url, headers=headers, allow_redirects=True, timeout=30)
        if r.status_code == 200 and len(r.content) > 100 and r.content[:2] == b"PK":
            with open(dest, "wb") as f:
                f.write(r.content)
            print(f"Скачано: {len(r.content)} байт")
            return True
        print(f"Скачивание: статус {r.status_code}, размер {len(r.content)}, первые байты: {r.content[:4]}")
    except Exception as e:
        print(f"Ошибка скачивания: {e}")
    return False


def download_xlsx(share_url: str, dest: str) -> bool:
    """Скачивает xlsx по ссылке OneDrive (несколько способов)."""
    if not HAS_REQUESTS:
        print("Библиотека requests недоступна, скачивание невозможно", file=sys.stderr)
        return False

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }

    # ================================================================
    # Способ 1: Извлечение FileGetUrl из HTML-страницы (САМЫЙ НАДЁЖНЫЙ)
    # ================================================================
    try:
        print(f"Открываем ссылку: {share_url}")
        r = requests.get(share_url, headers=headers, allow_redirects=True, timeout=20)
        page_text = r.text
        print(f"Страница загружена: {len(page_text)} символов, URL: {r.url[:80]}...")

        # 1a. Ищем FileGetUrl (содержит tempauth-токен для скачивания)
        file_get_match = re.search(
            r'"FileGetUrl"\s*:\s*"(https://[^"]+)"',
            page_text,
        )
        if file_get_match:
            dl_url = _unescape_url(file_get_match.group(1))
            print(f"Найден FileGetUrl (длина {len(dl_url)})")
            if _try_download(dl_url, dest, headers):
                return True
            print("FileGetUrl не дал валидный xlsx, пробуем другие способы...")

        # 1b. Ищем @content.downloadUrl
        content_dl_match = re.search(
            r'"@content\.downloadUrl"\s*:\s*"(https://[^"]+)"',
            page_text,
        )
        if content_dl_match:
            dl_url = _unescape_url(content_dl_match.group(1))
            print(f"Найден @content.downloadUrl (длина {len(dl_url)})")
            if _try_download(dl_url, dest, headers):
                return True
            print("@content.downloadUrl не дал валидный xlsx, пробуем другие способы...")

        # 1c. Ищем любой URL с tempauth на my.microsoftpersonalcontent.com
        tempauth_match = re.search(
            r'(https://my\.microsoftpersonalcontent\.com/[^"\s]+tempauth=[^"\s]+)',
            page_text,
        )
        if tempauth_match:
            dl_url = _unescape_url(tempauth_match.group(1))
            print(f"Найден tempauth URL на my.microsoftpersonalcontent.com (длина {len(dl_url)})")
            if _try_download(dl_url, dest, headers):
                return True

        # 1d. Ищем любой URL с tempauth на onedrive.live.com
        tempauth_match2 = re.search(
            r'(https://onedrive\.live\.com/[^"\s]+download\.aspx[^"\s]*tempauth=[^"\s]+)',
            page_text,
        )
        if tempauth_match2:
            dl_url = _unescape_url(tempauth_match2.group(1))
            print(f"Найден tempauth URL на onedrive.live.com (длина {len(dl_url)})")
            if _try_download(dl_url, dest, headers):
                return True

    except Exception as e:
        print(f"Ошибка при разборе страницы: {e}")

    # ================================================================
    # Способ 2: Graph API shares (может потребовать токен)
    # ================================================================
    try:
        import base64
        encoded = base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
        api_url = f"https://graph.microsoft.com/v1.0/shares/u!{encoded}/root/content"
        print(f"Пробуем Graph API: {api_url[:80]}...")
        r = requests.get(api_url, headers=headers, allow_redirects=True, timeout=30)
        if r.status_code == 200 and len(r.content) > 100 and r.content[:2] == b"PK":
            with open(dest, "wb") as f:
                f.write(r.content)
            print(f"Скачано через Graph API: {len(r.content)} байт")
            return True
        print(f"Graph API: статус {r.status_code}")
    except Exception as e:
        print(f"Graph API: ошибка — {e}")

    print("Не удалось скачать файл с OneDrive", file=sys.stderr)
    return False


def convert_xlsx_to_json(xlsx_path: str, json_path: str) -> bool:
    """Конвертирует xlsx в JSON."""
    wb = openpyxl.load_workbook(xlsx_path)
    all_data = {}

    for sheet_name, config in SHEETS_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"Лист «{sheet_name}» не найден, пропускаю", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h else "" for h in rows[0]]
        data_rows = []

        for row in rows[1:]:
            obj = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    header_name = headers[i]
                    cell_val = str(val) if val is not None else ""
                    # Конвертируем пути картинок и файлов в web-пути
                    if header_name == "Image":
                        cell_val = _convert_image_path(cell_val)
                    elif header_name == "Файл":
                        cell_val = _convert_file_path(cell_val)
                    obj[header_name] = cell_val
            data_rows.append(obj)

        all_data[config["id"]] = {
            "title": config["title"],
            "sheet": sheet_name,
            "headers": headers,
            "rows": data_rows,
            "total": len(data_rows),
        }
        print(f"  {sheet_name}: {len(data_rows)} строк")

    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    size = os.path.getsize(json_path)
    print(f"JSON сохранён: {json_path} ({size / 1024:.1f} КБ)")
    return True


def main():
    xlsx_path = LOCAL_XLSX

    # 1. Проверяем локальный xlsx в репозитории
    if os.path.isfile(xlsx_path):
        print(f"Найден локальный файл: {xlsx_path}")
    else:
        # 2. Пробуем скачать с OneDrive
        print("Локальный xlsx не найден, скачиваем с OneDrive...")
        xlsx_path = "/tmp/exam_tickets.xlsx"
        if not download_xlsx(ONEDRIVE_SHARE_URL, xlsx_path):
            print("ОШИБКА: не удалось получить xlsx файл", file=sys.stderr)
            sys.exit(1)

    # Проверяем что файл валидный
    try:
        with open(xlsx_path, "rb") as f:
            header = f.read(4)
        if header[:2] != b"PK":
            print(f"ОШИБКА: файл {xlsx_path} не является xlsx (не ZIP)", file=sys.stderr)
            sys.exit(1)
    except Exception as e:
        print(f"ОШИБКА: не удалось прочитать файл — {e}", file=sys.stderr)
        sys.exit(1)

    print("Конвертация xlsx → JSON...")
    if not convert_xlsx_to_json(xlsx_path, OUTPUT_PATH):
        sys.exit(1)

    print("Готово!")


if __name__ == "__main__":
    main()
